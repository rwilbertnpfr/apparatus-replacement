"""
build.py — NPFR Fleet & Equipment Replacement Dashboard build script
Reads the latest data/fleet_*.json as the single source of truth.
Run before git push:
    python build.py
Outputs:
    data/index.json
    exports/NPFR_Fleet_Replacement_YYYYMMDD.xlsx
"""
import json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

BASE = Path(__file__).parent
DATA = BASE / "data"
EXPORTS = BASE / "exports"
EXPORTS.mkdir(exist_ok=True)

# ── Brand palette — matches /splan build.py HEX dict ────────────────────────
HEX = {
    "bg": "1C0505", "edge": "7B1416", "edge_dark": "4A0A0C",
    "gold": "C8A040", "gold_lt": "E8C870", "cream": "F5E6C8",
    "subtitle": "A07850", "light_row": "FFF5E6", "alt_row": "FAEBD7",
}
CAT_COLOR = {"Engine": "B71C1C", "Aerial": "E65100", "Rescue": "6A1B9A", "Staff": "1565C0"}
TYPES = ["Engine", "Aerial", "Rescue", "Staff"]

def fill(h): return PatternFill("solid", fgColor=h)
def font(h="F5E6C8", bold=False, size=10, italic=False):
    return Font(color=h, bold=bold, size=size, italic=italic, name="Calibri")
def ctr(wrap=False): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def lft(wrap=False): return Alignment(horizontal="left", vertical="center", wrap_text=wrap, indent=1)
def side(s="thin", c=HEX["gold"]): return Side(style=s, color=c)
def bdr(): return Border(left=side(), right=side(), top=side(), bottom=side())
def sc(ws, row, col, val, fnt=None, fil=None, aln=None, numfmt=None):
    c = ws.cell(row=row, column=col, value=val)
    if fnt: c.font = fnt
    if fil: c.fill = fil
    if aln: c.alignment = aln
    if numfmt: c.number_format = numfmt
    return c
def mc(ws, r1, c1, r2, c2): ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def blend(hexcolor, with_hex="FFFFFF", factor=0.85):
    a = tuple(int(hexcolor[i:i+2], 16) for i in (0, 2, 4))
    b = tuple(int(with_hex[i:i+2], 16) for i in (0, 2, 4))
    out = tuple(round(a[i]*(1-factor) + b[i]*factor) for i in range(3))
    return "".join(f"{v:02X}" for v in out)

CAT_TINT = {c: blend(v, factor=0.85) for c, v in CAT_COLOR.items()}
CURRENCY = '$#,##0;($#,##0);"-"'
PCT = "0.0%"
N_FUTURE = 6

# ── Load latest fleet JSON ───────────────────────────────────────────────────
def load_plan():
    files = sorted(DATA.glob("fleet_*.json"), reverse=True)
    if not files:
        raise FileNotFoundError("No fleet_*.json files found in data/")
    with open(files[0]) as f:
        return json.load(f), files[0]

# ── Rebuild data/index.json ──────────────────────────────────────────────────
def rebuild_index():
    files = sorted(DATA.glob("fleet_*.json"), reverse=True)
    names = [f.name for f in files]
    with open(DATA / "index.json", "w") as f:
        json.dump({"files": names}, f, indent=2)
    return names

# ── Excel export ──────────────────────────────────────────────────────────────
def build_excel(plan, export_name):
    units = plan["units"]
    escalators = plan["escalators"]

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Data")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = HEX["gold"]

    widths = [10, 22, 14, 12, 14, 15, 15, 16, 12, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for col in range(1, 15):
        ws.cell(row=1, column=col).fill = fill(HEX["edge_dark"])
    ws.row_dimensions[1].height = 6

    ws.row_dimensions[2].height = 24
    mc(ws, 2, 1, 2, 10)
    sc(ws, 2, 1, "NPFR FLEET & EQUIPMENT — REPLACEMENT DATA",
       fnt=Font(name="Calibri", bold=True, size=14, color=HEX["edge_dark"]),
       fil=fill(HEX["gold"]), aln=ctr())

    headers = ["CID#", "UNIT IDENTIFIER", "VEHICLE TYPE", "ITEM TYPE",
               "IN-SERVICE YEAR", "ORIGINAL COST", "REPLACEMENT YEAR",
               "YEARS TO REPLACEMENT", "ESCALATOR %", "REPLACEMENT COST"]
    HEADER_ROW = 3
    for i, h in enumerate(headers, start=1):
        sc(ws, HEADER_ROW, i, h, fnt=Font(name="Calibri", bold=True, size=9, color=HEX["cream"]),
           fil=fill(HEX["edge"]), aln=ctr(wrap=True))
    ws.row_dimensions[HEADER_ROW].height = 30

    ws.column_dimensions["L"].width = 2
    ws.column_dimensions["M"].width = 12
    ws.column_dimensions["N"].width = 14
    ws.column_dimensions["O"].width = 16
    mc(ws, 2, 13, 2, 15)
    sc(ws, 2, 13, "ESCALATOR ASSUMPTIONS", fnt=Font(name="Calibri", bold=True, size=11, color=HEX["edge_dark"]),
       fil=fill(HEX["gold"]), aln=ctr())
    for i, h in enumerate(["VEHICLE TYPE", "VEHICLE ESC.", "EQUIPMENT ESC."], start=13):
        sc(ws, HEADER_ROW, i, h, fnt=Font(name="Calibri", bold=True, size=9, color=HEX["cream"]),
           fil=fill(HEX["edge"]), aln=ctr(wrap=True))
    ASSUMP_START = HEADER_ROW + 1
    for i, cat in enumerate(TYPES):
        r = ASSUMP_START + i
        bg = fill(HEX["alt_row"] if i % 2 else HEX["light_row"])
        sc(ws, r, 13, cat, fnt=Font(name="Calibri", bold=True, size=9, color=CAT_COLOR[cat]), fil=bg, aln=ctr())
        sc(ws, r, 14, escalators[cat]["vehicle"], fnt=font("0000FF", size=9), fil=bg, aln=ctr(), numfmt=PCT)
        sc(ws, r, 15, escalators[cat]["equipment"], fnt=font("0000FF", size=9), fil=bg, aln=ctr(), numfmt=PCT)
    ASSUMP_END = ASSUMP_START + len(TYPES) - 1
    TYPE_RANGE = f"$M${ASSUMP_START}:$M${ASSUMP_END}"
    VEH_RANGE = f"$N${ASSUMP_START}:$N${ASSUMP_END}"
    EQUIP_RANGE = f"$O${ASSUMP_START}:$O${ASSUMP_END}"

    dv_type = DataValidation(type="list", formula1=f'"{",".join(TYPES)}"', allow_blank=True)
    dv_item = DataValidation(type="list", formula1='"Vehicle,Equipment"', allow_blank=True)
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_item)

    def write_pair_formulas(v_row, e_row):
        for rr in (v_row, e_row):
            sc(ws, rr, 8, f"=IF(AND(ISNUMBER(E{rr}),ISNUMBER(G{rr})),G{rr}-E{rr},\"\")",
               fnt=font("000000", size=9), aln=ctr())
            sc(ws, rr, 9, f"=IFERROR(IF(D{rr}=\"Vehicle\",INDEX({VEH_RANGE},MATCH(C{rr},{TYPE_RANGE},0)),"
                          f"INDEX({EQUIP_RANGE},MATCH(C{rr},{TYPE_RANGE},0))),\"\")",
               fnt=font("000000", size=9), numfmt=PCT, aln=ctr())
            sc(ws, rr, 10, f"=IFERROR(F{rr}*(1+I{rr})^H{rr},\"\")",
               fnt=font("000000", bold=(rr == v_row), size=9), numfmt=CURRENCY)

    r = HEADER_ROW + 1
    for u in units:
        v_row, e_row = r, r + 1
        sc(ws, v_row, 1, u["cid"], fnt=font("0000FF", size=9))
        sc(ws, v_row, 2, u["unit"], fnt=font("0000FF", size=9))
        sc(ws, v_row, 3, u["type"], fnt=font("0000FF", bold=True, size=9))
        sc(ws, v_row, 4, "Vehicle", fnt=font("000000", size=9))
        sc(ws, v_row, 5, u["inServiceYear"], fnt=font("0000FF", size=9), aln=ctr())
        sc(ws, v_row, 6, u["originalCost"], fnt=font("0000FF", size=9), numfmt=CURRENCY)
        sc(ws, v_row, 7, u["replacementYear"], fnt=font("0000FF", size=9), aln=ctr())

        if isinstance(u["replacementYear"], (int, float)) and isinstance(u["inServiceYear"], (int, float)) \
                and u["replacementYear"] < u["inServiceYear"]:
            cell = ws.cell(row=v_row, column=7)
            cell.fill = fill("FFFF00")
            cell.comment = Comment(
                "Data flag: Replacement Year is earlier than In-Service Year. Verify this entry.", "build.py")

        sc(ws, e_row, 1, u["cid"], fnt=font("0000FF", size=9))
        sc(ws, e_row, 2, u["unit"], fnt=font("0000FF", size=9))
        sc(ws, e_row, 3, u["type"], fnt=font("0000FF", bold=True, size=9))
        sc(ws, e_row, 4, "Equipment", fnt=font("000000", italic=True, size=9))
        sc(ws, e_row, 5, u["inServiceYear"], fnt=font("0000FF", size=9), aln=ctr())
        sc(ws, e_row, 6, u["equipmentCost"], fnt=font("0000FF", size=9), numfmt=CURRENCY)
        sc(ws, e_row, 7, u["replacementYear"], fnt=font("0000FF", size=9), aln=ctr())

        write_pair_formulas(v_row, e_row)
        r += 2

    last_populated_row = r - 1
    for i in range(N_FUTURE):
        v_row, e_row = r, r + 1
        sc(ws, v_row, 4, "Vehicle", fnt=font("000000", italic=True, size=9))
        sc(ws, e_row, 4, "Equipment", fnt=font("000000", italic=True, size=9))
        write_pair_formulas(v_row, e_row)
        dv_type.add(f"C{v_row}")
        dv_item.add(f"D{v_row}:D{e_row}")
        r += 2
    table_last_row = r - 1
    dv_type.add(f"C{HEADER_ROW+1}:C{table_last_row}")

    for row_i in range(HEADER_ROW, table_last_row + 1):
        for col_i in range(1, 11):
            ws.cell(row=row_i, column=col_i).border = bdr()

    tbl = Table(displayName="FleetData", ref=f"A{HEADER_ROW}:J{table_last_row}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True,
                                         showFirstColumn=False, showLastColumn=False, showColumnStripes=False)
    ws.add_table(tbl)

    for cat in TYPES:
        ws.conditional_formatting.add(
            f"C{HEADER_ROW+1}:C{table_last_row}",
            CellIsRule(operator="equal", formula=[f'"{cat}"'], fill=fill(CAT_TINT[cat])))
        ws.conditional_formatting.add(
            f"A{HEADER_ROW+1}:J{table_last_row}",
            FormulaRule(formula=[f'$C{HEADER_ROW+1}="{cat}"'], fill=fill(CAT_TINT[cat])))

    ws.freeze_panes = f"A{HEADER_ROW+1}"

    out_path = EXPORTS / export_name
    wb.save(out_path)
    return out_path

# ── Console summary ──────────────────────────────────────────────────────────
def print_summary(plan):
    print("\n[4] Fleet summary by vehicle type")
    for t in TYPES:
        units = [u for u in plan["units"] if u["type"] == t]
        esc = plan["escalators"][t]
        total = 0
        for u in units:
            years = u["replacementYear"] - u["inServiceYear"]
            total += u["originalCost"] * (1 + esc["vehicle"]) ** years
            total += u["equipmentCost"] * (1 + esc["equipment"]) ** years
        print(f"  {t:<8} {len(units):>3} units   ${total:>14,.0f} projected replacement cost")

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("NPFR Fleet & Equipment Replacement Dashboard — build.py")
    print("=" * 60)

    print("\n[1] Rebuilding data/index.json")
    names = rebuild_index()
    print(f"  index.json → {len(names)} file(s), latest: {names[0] if names else 'none'}")

    print("\n[2] Loading fleet data")
    plan, src_file = load_plan()
    print(f"  Source: {src_file.name}")
    print(f"  Units: {len(plan['units'])}")

    print("\n[3] Building Excel workbook")
    stamp = datetime.now().strftime("%Y%m%d")
    export_name = f"NPFR_Fleet_Replacement_{stamp}.xlsx"
    build_excel(plan, export_name)
    print(f"  Saved: exports/{export_name}")

    if plan.get("exportFile") != export_name:
        plan["exportFile"] = export_name
        with open(src_file, "w") as f:
            json.dump(plan, f, indent=2)
        print(f"  Updated {src_file.name} with exportFile reference")

    print_summary(plan)
    print("\nDone.")

if __name__ == "__main__":
    main()
